""" Script for webscraping icecream flavours
"""
import bs4
import requests
import openpyxl
from openpyxl.styles import PatternFill

BASE_URL="https://gelatomessina.com"

def main():
    dietary_requirements=[
        ('/Vegan', 'icecream_vegan'),
        ('/Alcohol-Free', 'icecream_alcohol_free'),
        ('/Egg-Free', 'icecream_egg_free'),
        ('/Gluten-Free', 'icecream_gluten_free'),
        ('/Nut-Free', 'icecream_nut_free'),
    ]

    all_flavours=get_flavours_through_pagination(BASE_URL+'/collections/classic-flavours',add_description=True)
    print(all_flavours)

    wb=openpyxl.Workbook()
    ws = wb.active
    header_flavour=ws.cell(row=1, column=1, value="Flavour")
    header_flavour.fill = PatternFill(start_color="729fcf", fill_type="solid")
    header_description=ws.cell(row=1, column=2, value="Description")
    header_description.fill = PatternFill(start_color="729fcf", fill_type="solid")

    row=2
    for flavour in all_flavours:
        ws.cell(row=row,column=1).value=flavour[0]
        ws.cell(row=row,column=2).value=flavour[1]
        row+=1
    
    col=3
    for dietr in dietary_requirements:
        header_flavour=ws.cell(row=1, column=col, value=dietr[0][1:])
        header_flavour.fill = PatternFill(start_color="b4c7dc", fill_type="solid")

        comply_requirements=get_flavours_through_pagination(BASE_URL+'/collections/classic-flavours'+dietr[0])
        for row in ws.iter_rows(min_row=1, max_col=1):
            for cell in row:
                this_flavour=cell.value
                print(this_flavour)
                if this_flavour in comply_requirements:
                    ws.cell(row=cell.row,column=col).value="X"
        col+=1
    wb.save('flavours.xlsx')

def export_to_file(file_name, flavours):
    with open(file_name+'.txt', 'w') as f:
        for flavour in flavours:
            f.write(flavour)
            f.write("\n")    

def get_flavours_through_pagination(full_url,add_description=False):
    next=(True, full_url)
    flavours_collected=[]
    while True:
        full_url=next[1]
        response =requests.get(full_url)
        flavours=get_flavours(response,add_description=add_description)
        flavours_collected.extend(flavours)
        next=get_next_pagination_href(full_url)
        print("next is ", next)
        if next[0] == False:
            print("finished!")
            break
    return flavours_collected


def get_next_pagination_href(base_url):
    """ Returns a tuple with a boolean and the href if present

    If boolean True -> there is a next page, and the second element is the link

    """
    response =requests.get(base_url)
    soup = bs4.BeautifulSoup(response.content,"lxml")
    ul=soup.find('ul',{"class":"pagination flex items-center"})  
    if ul is not None:
        for element in ul.children:
            try:
                anchor=element.find("a")
                span=anchor.find("span")
                if "Next" in span.text:
                    return (True, BASE_URL+anchor['href']) 
            except:
                pass
    return (False, None)


def get_flavours(response,add_description=False):
    """Extracts the flavour names from a page and returns them in a list
    """
    soup = bs4.BeautifulSoup(response.content,"lxml")
    anchors=soup.find_all('a',{"class":"increase-target"})
    flavours=[]
    for anchor in anchors:
        span=anchor.find("span")
        name=str(span.text)
        name=name.replace("*","")        
        if add_description:
            description=get_flavour_description(BASE_URL+anchor['href'])
            flavour_and_desc=[name,description]
            flavours.append(flavour_and_desc)
        else:
            flavours.append(name)
    return flavours

def get_flavour_description(full_url):
    response =requests.get(full_url)
    soup = bs4.BeautifulSoup(response.content,"lxml")
    try:
        description=soup.find('span',{"data-mce-fragment":"1"}).text 
    except:
        try:
            description=soup.find('div',{"class":"flavours-description"}).text
        except:
            description=""
    print(full_url)
    print(description) 
    return description

if __name__ == "__main__":
    main()